"""
excel 工具：对 Excel 工作簿进行读取、写入与格式化操作。
params 中 action 字段决定执行的具体操作：

  read_sheet   读取工作表数据
  write_cells  写入单元格（支持批量）
  create       新建工作簿
  add_sheet    新增工作表
  delete_sheet 删除工作表
  list_sheets  列出所有工作表名
  set_style    设置单元格样式（字体 / 颜色 / 边框）
  to_csv       导出工作表为 CSV
  merge_cells  合并单元格
  freeze       冻结行或列

公共参数：
  path (str, 必填) 工作簿绝对路径（须在允许根目录内）
  sheet (str, 可选) 目标工作表名，默认取第一张
"""
import base64
import csv
import zipfile
from pathlib import Path
from typing import Any

from tools.sandbox import resolve_safe_path  # H1：路径沙箱

DEFAULT_READ_LIMIT_ROWS = 10_000
# 多模态：内嵌图片抽取护栏,防超大工作簿把请求体撑爆(base64 会随请求发往厂商)
MAX_XLSX_IMAGES = 50
MAX_XLSX_IMAGE_SIZE = 8 * 1024 * 1024  # 单图 8 MB

# xl/media 里的位图后缀 → media_type;矢量(emf/wmf)模型看不了,直接跳过不返回
_IMAGE_MEDIA_TYPES = {
    "png": "image/png",
    "jpg": "image/jpeg",
    "jpeg": "image/jpeg",
    "gif": "image/gif",
    "bmp": "image/bmp",
    "webp": "image/webp",
}

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


def _require_openpyxl():
    if not _HAS_OPENPYXL:
        raise ImportError("openpyxl 未安装，请运行 pip install openpyxl")


def _get_ws(wb, sheet_name: str | None):
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"工作表不存在: {sheet_name}")
        return wb[sheet_name]
    return wb.active


def _read_rows(ws, limit_rows: int) -> list[list]:
    rows: list[list] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
        if len(rows) >= limit_rows:
            break
    return rows


def _extract_comments(
    wb,
    sheet_name: str | None,
    limit_rows: int | None = None,
) -> list[dict]:
    """遍历目标工作表(无 sheet 则全部)的单元格批注。openpyxl 的 cell.comment 可靠。
    limit_rows 限制扫描行数(与 _read_rows 对齐),None=不限制。"""
    sheets = [_get_ws(wb, sheet_name)] if sheet_name else list(wb.worksheets)
    out: list[dict] = []
    for ws in sheets:
        max_row = limit_rows if limit_rows and limit_rows > 0 else None
        for row in ws.iter_rows(max_row=max_row):
            for cell in row:
                c = cell.comment
                if c is None:
                    continue
                out.append({
                    "sheet": ws.title,
                    "cell": cell.coordinate,
                    "text": (c.text or "").strip(),
                    "author": c.author or "",
                })
    return out


def _extract_images(path: Path) -> tuple[list[dict], bool]:
    """把 .xlsx 当 zip 直接读 xl/media/*(比 openpyxl 的 ws._images 稳),返回 base64 图片列表。
    返回 (images, truncated);超数量/超单图上限时截断。"""
    images: list[dict] = []
    truncated = False
    with zipfile.ZipFile(path) as z:
        media = sorted(n for n in z.namelist() if n.startswith("xl/media/"))
        for name in media:
            ext = name.rsplit(".", 1)[-1].lower() if "." in name else ""
            media_type = _IMAGE_MEDIA_TYPES.get(ext)
            if not media_type:
                continue  # 矢量/未知位图,模型无法识别,跳过
            if len(images) >= MAX_XLSX_IMAGES:
                truncated = True
                break
            data = z.read(name)
            if len(data) > MAX_XLSX_IMAGE_SIZE:
                continue  # 单图过大,跳过
            images.append({
                "media_type": media_type,
                "data_b64": base64.b64encode(data).decode("ascii"),
            })
    return images, truncated


async def execute(params: dict[str, Any]) -> Any:
    _require_openpyxl()

    path_str: str = params.get("path", "")
    if not path_str:
        raise ValueError("缺少必填参数 path")

    # H1：路径沙箱校验
    allow_outside_roots: bool = params.get("allow_outside_roots", False)
    path = resolve_safe_path(path_str, allow_outside_roots=allow_outside_roots)
    action: str = params.get("action", "read_sheet")
    sheet_name: str | None = params.get("sheet")

    # ── 新建工作簿 ─────────────────────────────────────────
    if action == "create":
        title: str = params.get("title", "Sheet1")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        return {"created": str(path), "sheet": title}

    if action == "create_table":
        sheet_title: str = params.get("sheet", "输出")
        headers: list[Any] = params.get("headers", [])
        rows: list[list[Any]] = params.get("rows", [])

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_title[:31] or "输出"

        row_index = 1
        if headers:
            for col_index, value in enumerate(headers, start=1):
                cell = ws.cell(row=row_index, column=col_index, value=value)
                cell.font = Font(bold=True, color="2F3437")
                cell.fill = PatternFill(fill_type="solid", fgColor="E9E2D6")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            row_index += 1

        for row in rows:
            values = row if isinstance(row, list) else [row]
            for col_index, value in enumerate(values, start=1):
                cell = ws.cell(row=row_index, column=col_index, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            row_index += 1

        ws.freeze_panes = "A2" if headers else None
        row_lengths = [len(row) if isinstance(row, list) else 1 for row in rows]
        max_col = max([len(headers), *row_lengths, 1])
        for col_index in range(1, max_col + 1):
            letter = get_column_letter(col_index)
            width = 12
            for cell in ws[letter]:
                width = max(width, min(len(str(cell.value or "")) + 2, 42))
            ws.column_dimensions[letter].width = width

        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        return {"path": str(path), "sheet": ws.title, "rows": len(rows)}

    if not path.exists():
        raise FileNotFoundError(f"文件不存在: {path}")

    wb = openpyxl.load_workbook(path)

    # ── 列出工作表 ─────────────────────────────────────────
    if action == "list_sheets":
        return {"sheets": wb.sheetnames}

    # ── 新增工作表 ─────────────────────────────────────────
    if action == "add_sheet":
        name: str = params.get("name", "新工作表")
        if name in wb.sheetnames:
            raise ValueError(f"工作表已存在: {name}")
        wb.create_sheet(title=name)
        wb.save(path)
        return {"added": name}

    # ── 删除工作表 ─────────────────────────────────────────
    if action == "delete_sheet":
        ws = _get_ws(wb, sheet_name)
        del wb[ws.title]
        wb.save(path)
        return {"deleted": ws.title}

    # ── 读取工作表数据 ──────────────────────────────────────
    if action == "read_sheet":
        ws = _get_ws(wb, sheet_name)
        limit_rows = int(params.get("limit_rows", DEFAULT_READ_LIMIT_ROWS))
        if limit_rows <= 0 or limit_rows > DEFAULT_READ_LIMIT_ROWS:
            limit_rows = DEFAULT_READ_LIMIT_ROWS
        rows = _read_rows(ws, limit_rows)
        return {
            "sheet": ws.title,
            "rows": rows,
            "row_count": len(rows),
            "col_count": ws.max_column,
            "truncated": ws.max_row > len(rows),
        }

    # ── 读取单元格批注 ─────────────────────────────────────
    if action == "comments":
        limit_rows = int(params.get("limit_rows", DEFAULT_READ_LIMIT_ROWS))
        if limit_rows <= 0 or limit_rows > DEFAULT_READ_LIMIT_ROWS:
            limit_rows = DEFAULT_READ_LIMIT_ROWS
        comments = _extract_comments(wb, sheet_name, limit_rows)
        return {"comments": comments, "count": len(comments)}

    # ── 抽取内嵌图片(zip 读 xl/media,返回 base64)─────────────
    if action == "images":
        images, truncated = _extract_images(path)
        result = {"images": images, "count": len(images)}
        if truncated:
            result["truncated"] = True
        return result

    # ── 富读取:一次返回正文 + 批注 + 内嵌图片(多模态链路用)────
    if action == "read_rich":
        ws = _get_ws(wb, sheet_name)
        limit_rows = int(params.get("limit_rows", DEFAULT_READ_LIMIT_ROWS))
        if limit_rows <= 0 or limit_rows > DEFAULT_READ_LIMIT_ROWS:
            limit_rows = DEFAULT_READ_LIMIT_ROWS
        rows = _read_rows(ws, limit_rows)
        # comments 传 ws.title 而非 sheet_name: sheet_name=None 时 rows 取 active 表,
        # comments 须同表对齐(否则会扫全部 worksheet 的批注却只配 active 表的 rows)。
        comments = _extract_comments(wb, ws.title, limit_rows)
        images, images_truncated = _extract_images(path)
        return {
            "sheet": ws.title,
            "rows": rows,
            "row_count": len(rows),
            "col_count": ws.max_column,
            "truncated": ws.max_row > len(rows),
            "comments": comments,
            "images": images,
            "images_truncated": images_truncated,
        }

    # ── 写入单元格 ─────────────────────────────────────────
    if action == "write_cells":
        ws = _get_ws(wb, sheet_name)
        cells: list[dict] = params.get("cells", [])
        for item in cells:
            ws[item["cell"]] = item.get("value")
        wb.save(path)
        return {"written": len(cells), "sheet": ws.title}

    # ── 设置单元格样式 ─────────────────────────────────────
    if action == "set_style":
        ws = _get_ws(wb, sheet_name)
        cell_ref: str = params.get("cell", "A1")
        cell = ws[cell_ref]
        style: dict = params.get("style", {})
        if "bold" in style or "font_size" in style or "font_color" in style:
            cell.font = Font(
                bold=style.get("bold", False),
                size=style.get("font_size"),
                color=style.get("font_color"),
            )
        if "bg_color" in style:
            cell.fill = PatternFill(fill_type="solid", fgColor=style["bg_color"])
        if "align" in style:
            cell.alignment = Alignment(horizontal=style["align"])
        wb.save(path)
        return {"styled": cell_ref}

    # ── 导出 CSV（H1：out_path 同样校验）─────────────────────
    if action == "to_csv":
        ws = _get_ws(wb, sheet_name)
        out_path_str: str = params.get("out_path", str(path.with_suffix(".csv")))
        out_path = resolve_safe_path(out_path_str)  # H1 + M5.5：out_path 也需沙箱
        encoding: str = params.get("encoding", "utf-8-sig")
        with open(out_path, "w", newline="", encoding=encoding) as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)
        return {"csv_path": str(out_path), "rows": ws.max_row}

    # ── 合并单元格 ─────────────────────────────────────────
    if action == "merge_cells":
        ws = _get_ws(wb, sheet_name)
        cell_range: str = params.get("range", "A1:B1")
        ws.merge_cells(cell_range)
        wb.save(path)
        return {"merged": cell_range}

    # ── 冻结行/列 ──────────────────────────────────────────
    if action == "freeze":
        ws = _get_ws(wb, sheet_name)
        freeze_at: str = params.get("at", "A2")
        ws.freeze_panes = freeze_at
        wb.save(path)
        return {"freeze_panes": freeze_at}

    raise ValueError(f"未知 action: {action}")
